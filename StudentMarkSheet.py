import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

#Read Input Excel File
#importing static data file and performing operations on it
df=pd.read_excel("studentsdata.input.xlsx")

subject_order={"Chemistry":0,"Physics":1,"English":2}
df["SubjectOrder"]=df["Subject"].map(subject_order)
# Sort data by 'Name' to group student marks together
df=df.sort_values(by=["Name","Subject"]).reset_index(drop=True)

#converts the 'Date' column from a string format to a proper Datetime object in pandas
df["D.O.B"]=pd.to_datetime(df["D.O.B"], dayfirst=True).dt.strftime("%b %d, %Y")
#Format datetime to "Month Date, Yeat" string

grouped=df.groupby("Name")
final_rows=[]
sr_no=1

for name,group in grouped:
    group=group.reset_index(drop=True)
    perc=(group["Marks"].sum()/240)*100
# for name in df["Name"].unique():
#     student_df=df[df["Name"]==name]
#     date=student_df.iloc[0]["Date"]

#     marks_dict={"Chemistry":None, "Physics": None, "English": None}
#     for _, row in student_df.iterrows():
#         marks_dict[row["Subject"]]=row["Marks"]

#Adds remarks based on Percentage
    if perc > 90:
        grade="A"
        remarks= "Excellent"
    elif 80 < perc < 90:
        grade="B"
        remarks= "Very Good"
    elif 70 < perc < 80:
        grade="C"
        remarks= "Good"
    elif 50 < perc < 70:
        grade="D"
        remarks= "Average"
    elif 40 < perc < 50:
        grade="E"
        remarks= "Pass"
    else:
        grade="F"
        remarks= "Needs Improvement"
    

    # final_rows.append({"Name": name, "Date": formatted_date,"Label":"Chemistry","Value":marks_dict["Chemistry"]})
    # final_rows.append({"Name":"", "Date": "","Label":"Physics","Value":marks_dict["Physics"]})
    # final_rows.append({"Name":"", "Date": "","Label":"English","Value":marks_dict["English"]})
    # final_rows.append({"Name": "", "Date": "","Label":"Total","Value":total})
    # final_rows.append({"Name": "", "Date": "","Label":"Percentage","Value":perc})
    # final_rows.append({"Name": "", "Date": "","Label":"Remarks","Value":remarks})
    dob_value=group.loc[0,"D.O.B"]
    for i in range(len(group)):
        row={
            "Sr.No": sr_no if i==0 else"",
            "Student Name": name if i==0 else"",
            "D.O.B": dob_value if i==0 else"",
            "Subject": group.loc[i,"Subject"],
            "Marks": group.loc[i,"Marks"],
            "Total": group.loc[i,"Total"],
            "Grade": grade if i==2 else"",
            "Remarks": remarks if i==2 else"",
        }
        final_rows.append(row)
    sr_no+=1    

#Groups all marks by students name and Calculate total marks for each student
# total_marks=df_sorted.groupby(["Name","Date"])["Marks"].sum().reset_index()
#Renames "Marks" to "Total" for clarity
# total_marks.rename(columns={"Marks": "Total"}, inplace=True)

#calculate percentage rounded upto 2 decimal places
# total_marks["Percentage"]=round((total_marks["Total"]/240)*100,2)
# total_marks["Remarks"]=total_marks["Percentage"].apply(student_remarks)

#df_sorted:contains sttudent wise marks in proper oder
#total_marks

# df_final=pd.merge(df_sorted, total_marks, on=["Name","Date"],how="left")
final_df=pd.DataFrame(final_rows)
output_file3="studentMarkSheet.xlsx"
final_df.to_excel(output_file3, index= False, engine="openpyxl")

# Create Excel workbook
workbook = load_workbook(output_file3)
sheet = workbook.active
sheet.title="Student"

# Merge cells
headers=[cell.value for cell in sheet[1]]
print("Headers:", headers)

col_idx={name:headers.index(name)+1 for name in headers}
max_row=sheet.max_row
row=2
while row<=max_row:
    start_row=row
    srno=sheet.cell(row=row,column=col_idx["Sr.No"]).value
    name=sheet.cell(row=row,column= col_idx["Student Name"]).value
    dob=sheet.cell(row=row,column= col_idx["D.O.B"]).value
    grade=sheet.cell(row=row,column=col_idx["Grade"]).value
    remarks=sheet.cell(row=row,column=col_idx["Remarks"]).value

    count=1
    while (row+count<= max_row and 
           sheet.cell(row+count,column=col_idx["Sr.No"]).value=="" and
           sheet.cell(row+count,column=col_idx["Student Name"]).value=="" and
           sheet.cell(row+count,column=col_idx["D.O.B"]).value=="" and
           sheet.cell(row+count,column=col_idx["Grade"]).value=="" and
           sheet.cell(row+count,column=col_idx["Remarks"]).value==""): 
        count+=1
 #   str(sheet.cell(row+count,column=2).value).strip()==str(name).strip() and
        #   str(sheet.cell(row+count,column=3).value).strip()==str(date).strip() and
    # if count>1:
    #     sheet.merge_cells(start_row=row,start_column=1,end_row=row+count-1,end_column=1)  
    #     sheet.merge_cells(start_row=row,start_column=2,end_row=row+count-1,end_column=2)   
    #     sheet.merge_cells(start_row=row,start_column=3,end_row=row+count-1,end_column=3)
    #     sheet.merge_cells(start_row=row,start_column=6,end_row=row+count-1,end_column=6)
    #     sheet.merge_cells(start_row=row,start_column=7,end_row=row+count-1,end_column=7)

    if count >1:
        for col in ["Sr.No","Student Name","D.O.B","Grade","Remarks"]:
            sheet.merge_cells(start_row=start_row,start_column=col_idx[col_name], end_row=start_row+count-1, end_column=col_idx[col_name])
        # cell=sheet.cell(row=row,column=col)
            cell.alignment= Alignment(horizontal="center",vertical="center")
            print(f"Merged {col_name} from row{start_row+count -1}")
    row+=count


grade_col=[cell.value for cell in sheet[1]].index("Grade")+1

red_background = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
yellow_background = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


for row in sheet.iter_rows(min_row=2,min_col=grade_col,max_col=grade_col):
        for cell in row:
            if cell.value=="F":
                cell.fill= red_background
            elif cell.value=="E":
                cell.fill= yellow_background

# Save workbook
workbook.save(output_file3)
print("Excel file created successfully")